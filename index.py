# -*- coding: utf-8 -*-
"""newintern3.py-checkpoint.ipynb

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/1TOQ19moNwFj6rLqo7VBL_Wnn0fbUZR3I
"""

import pandas as pd
import re
from openpyxl.styles import PatternFill
from openpyxl import Workbook

try:
    df = pd.read_excel('transportmaster.xls')  # Corrected to read from an Excel file
    transporter_code_pattern = re.compile(r'^[A-Z0-9]{6}$')
    transporter_name_pattern = re.compile(r'^.+')
    special_characters_pattern = re.compile(r'[@#<>!$%&*]')
    transporter_contact_pattern = re.compile(r'^[A-Za-z\s]+$')
    company_pincode_pattern = re.compile(r'^[0-9]{6}$')
    company_gst_pattern = re.compile(r'^\d{2}[A-Z]{5}\d{4}[A-Z]{1}[1-9A-Z]{1}Z\d[0-9A-Z]{1}$')
    company_pan_pattern = re.compile(r'^[A-Z]{5}\d{4}[A-Z]{1}$')
    vendor_type_pattern = re.compile(r'^(owned|transporter)$', re.IGNORECASE)

    error_cells = []

    # Check transporterCode (unique and pattern)
    seen_transporter_codes = set()
    for index, row in df.iterrows():
        if not pd.isna(row['transporterCode']):
            transporter_code = str(row['transporterCode']).strip()
            if transporter_code in seen_transporter_codes or not transporter_code_pattern.match(transporter_code):
                error_cells.append((index, 'transporterCode'))
            else:
                seen_transporter_codes.add(transporter_code)

        # Check transporterName (required and pattern)
        if pd.isna(row['transporterName']) or not transporter_name_pattern.match(str(row['transporterName']).strip()):
            error_cells.append((index, 'transporterName'))

        # Check transporterContactPerson (no special characters)
        if special_characters_pattern.search(str(row['transporterContactPerson'])):
            error_cells.append((index, 'transporterContactPerson'))

        # Check transporterContactPersonNumber (10 digits only)
        if not pd.isna(row['transporterContactPersonNumber']):
            contact_number = str(row['transporterContactPersonNumber']).strip()
            if not (len(contact_number) == 10 and contact_number.isdigit()):
                error_cells.append((index, 'transporterContactPersonNumber'))

        # Check companyCity (no special characters)
        if special_characters_pattern.search(str(row['companyCity'])):
            error_cells.append((index, 'companyCity'))

        # Check companyPincode (6 digits only)
        if not company_pincode_pattern.match(str(row['companyPincode']).strip()):
            error_cells.append((index, 'companyPincode'))

        # Check companyGst (Indian GST format)
        if not pd.isna(row['companyGst']):  # Check for NaN values
            if not company_gst_pattern.match(str(row['companyGst']).strip()):
                error_cells.append((index, 'companyGst'))

        # Check companyPan (Indian PAN format)
        if not pd.isna(row['companyPan']):  # Check for NaN values
            if not company_pan_pattern.match(str(row['companyPan']).strip()):
                error_cells.append((index, 'companyPan'))

        # Check vendorType (owned or transporter)
        if not vendor_type_pattern.match(str(row['vendorType']).strip()):
            error_cells.append((index, 'vendorType'))

    # Create a new Excel file and write the data to it
    wb = Workbook()
    ws = wb.active
    ws.title = 'Error Report'

    # Write the headers to the Excel file
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx).value = col_name

    # Write the data to the Excel file
    for r, row in df.iterrows():
        for c, value in enumerate(row, 1):
            ws.cell(row=r+2, column=c).value = value  # Start writing data from the second row onwards

    # Color the error cells red
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    for error_cell in error_cells:
        row, col_name = error_cell
        col_idx = df.columns.get_loc(col_name) + 1
        ws.cell(row=row+2, column=col_idx).fill = red_fill  # Adjust row index to match Excel rows

    # Save the Excel file
    wb.save('output3.xlsx')

except FileNotFoundError:
    print("The file 'transportmaster.xls' was not found. Please make sure the file exists in the same directory.")
except Exception as e:
    print("An error occurred:", str(e))

from google.colab import drive
drive.mount('/content/drive')

